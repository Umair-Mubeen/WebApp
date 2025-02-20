import re
import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment


def extract_pay_and_designation_from_pdf(pdf_path, personnel_excel):
    try:
        # Load personnel Excel file
        df = pd.read_excel(personnel_excel)

        # Add missing columns if not present
        for column in ['Basic Pay', 'PayRoll Designation']:
            if column not in df.columns:
                df[column] = None

        # Clean personnel numbers
        personnel_list = df['Personnel'].dropna().astype(str).tolist()
        personnel_list = [
            str(int(float(num))) for num in personnel_list if num.strip() and num.strip() != "0"
        ]

        if not personnel_list:
            print("No valid personnel found.")
            return

        # Open the PDF
        with pdfplumber.open(pdf_path) as pdf:
            for idx, personnel in enumerate(personnel_list):
                print(f"\nSearching for Personnel # {personnel}")

                basic_pay, designation = None, None
                found_pers = False

                # Search for personnel number in each page
                for page in pdf.pages:
                    text = page.extract_text()

                    if f"Personnel Number: {personnel}" in text:
                        found_pers = True

                        # Extract Basic Pay
                        pay_pattern = r'Basic Pay\s+([\d,]+\.\d{2})'
                        pay_match = re.search(pay_pattern, text)
                        if pay_match:
                            basic_pay = pay_match.group(1)

                        # Extract Designation
                        desig_pattern = r'Designation\s*:\s*([A-Za-z\s]+)'
                        desig_match = re.search(desig_pattern, text)
                        if desig_match:
                            designation = desig_match.group(1).strip()

                        break  # Stop searching once found

                # Update DataFrame
                df.at[idx, 'Basic Pay'] = basic_pay if basic_pay else 'Not In Pay Roll'
                df.at[idx, 'PayRoll Designation'] = designation if designation else 'Unknown'

                print(f'Personnel # {personnel} -> Basic Pay: {df.at[idx, "Basic Pay"]}, Designation: {df.at[idx, "PayRoll Designation"]}')

        # Save updated DataFrame to Excel
        df.to_excel(personnel_excel, index=False)

        # Format the Excel file
        format_excel(personnel_excel, df)

    except Exception as e:
        print(f'An error occurred: {str(e)}')


def format_excel(personnel_excel, df):
    try:
        workbook = load_workbook(personnel_excel)
        sheet = workbook.active

        # Define styles
        border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
        fill_style_header = PatternFill(start_color='FFCCFF', end_color='FFCCFF', fill_type='solid')  # Light purple
        fill_style_missing = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # Light yellow
        fill_style_row = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')  # Light gray for rows
        font_style = Font(name='Book Antiqua', bold=True, color='000000')  # Black text
        alignment_style = Alignment(wrap_text=True)

        # Apply styles to header row
        for col in range(1, len(df.columns) + 1):
            header_cell = sheet.cell(row=1, column=col)
            header_cell.fill = fill_style_header
            header_cell.font = font_style
            header_cell.border = border_style
            header_cell.alignment = alignment_style

        # Apply styles to data rows
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = border_style
                cell.alignment = alignment_style

                # Fill entire row up to "Basic Pay" with background color
                if col < df.columns.get_loc('Basic Pay') + 1:
                    cell.fill = fill_style_row

                # Highlight missing Basic Pay or Designation
                if col == df.columns.get_loc('Basic Pay') + 1 and cell.value == 'Not In Pay Roll':
                    cell.fill = fill_style_missing
                if col == df.columns.get_loc('PayRoll Designation') + 1 and cell.value == 'Unknown':
                    cell.fill = fill_style_missing

        # Adjust column widths for better readability
        for col in range(1, len(df.columns) + 1):
            column_letter = sheet.cell(row=1, column=col).column_letter
            max_length = max(len(str(cell.value)) for cell in sheet[column_letter])
            sheet.column_dimensions[column_letter].width = max_length + 2  # Adding padding

        # Save the formatted workbook
        workbook.save(personnel_excel)

    except Exception as e:
        print(f'Error formatting Excel: {str(e)}')


# Example usage
pdf_file_path = 'C:/Users/ACS/Downloads/JAN-25.pdf'
personnel_excel_path = 'C:/Users/ACS/Pictures/Disposition.xlsx'
extract_pay_and_designation_from_pdf(pdf_file_path, personnel_excel_path)
