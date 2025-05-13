import re
import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

import pandas as pd
import pdfplumber
import re
from openpyxl import load_workbook


def extract_and_append_missing_personnel(pdf_path, personnel_excel):
    try:
        # Load personnel Excel file
        df = pd.read_excel(personnel_excel)

        # Ensure required columns exist
        for column in ['Personnel', 'Name', 'Designation', 'Basic Pay']:
            if column not in df.columns:
                df[column] = None

        # Clean personnel numbers in Excel
        personnel_list = df['Personnel'].dropna().astype(str).tolist()
        personnel_list = [
            str(int(float(num))) for num in personnel_list if num.strip() and num.strip() != "0"
        ]

        # Open the PDF
        missing_data = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()

                # Extract all personnel numbers
                personnel_matches = re.findall(r'Personnel Number:\s*(\d+)', text)
                for personnel in personnel_matches:
                    if personnel not in personnel_list:  # If personnel is missing in Excel
                        print(f"Missing Personnel Found: {personnel}")

                        # Extract Name
                        name_match = re.search(r'Name:\s*([A-Za-z\s]+)', text)
                        name = name_match.group(1).strip() if name_match else "Unknown"

                        # Extract Basic Pay
                        pay_match = re.search(r'Basic Pay\s+([\d,]+\.\d{2})', text)
                        basic_pay = pay_match.group(1) if pay_match else "Not In Pay Roll"

                        # Extract Designation
                        desig_match = re.search(r'Designation\s*:\s*([A-Za-z\s]+)', text)
                        designation = desig_match.group(1).strip() if desig_match else "Unknown"

                        # Store missing personnel
                        missing_data.append([personnel, name, designation, basic_pay])

        # If missing personnel found, append to Excel
        if missing_data:
            df_missing = pd.DataFrame(missing_data, columns=['Personnel', 'Name', 'Designation', 'Basic Pay'])

            # Append missing personnel to the existing DataFrame
            df = pd.concat([df, df_missing], ignore_index=True)

            # Save updated DataFrame back to Excel
            df.to_excel(personnel_excel, index=False)

            print("Missing personnel details appended to the existing Excel sheet.")

        else:
            print("No missing personnel found.")

    except Exception as e:
        print(f'An error occurred: {str(e)}')


# Example usage
pdf_file_path = 'C:/Users/ACS/Videos/Mar-2025.pdf'
personnel_excel_path = 'C:/Users/ACS/Videos/March PayRoll.xlsx'
extract_and_append_missing_personnel(pdf_file_path, personnel_excel_path)
