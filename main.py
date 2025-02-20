import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment


def extract_basic_pay(path, personnel_excel):
    try:
        # Read the personnel numbers from the provided Excel file
        df = pd.read_excel(personnel_excel)

        # Add a new column 'Basic Pay' to store the extracted basic pay, if it doesn't already exist
        if 'Basic Pay' not in df.columns:
            df['Basic Pay'] = None

        # Clean the personnel list, ignoring empty and zero values
        personnel_list = df['Personnel'].dropna().astype(str).tolist()
        personnel_list = [
            str(int(float(num))) for num in personnel_list
            if num.strip() and num.strip() != "0"
        ]

        if not personnel_list:
            print("No valid personnel found.")
            return

        with open(path, 'r', encoding='utf-8', errors='replace') as file:
            lines = file.readlines()

        for idx, personnel in enumerate(personnel_list):
            print(f"\nSearching for Personnel # {personnel}")

            personnel = personnel.strip()
            basic_pay = None
            found_pers = False
            pers_column = None

            for line in lines:
                left_column = line[:80].strip()
                right_column = line[80:].strip()

                # Check if personnel number is found
                if f"Pers #: {personnel}" in left_column:
                    found_pers = True
                    pers_column = 'left'
                    continue
                elif f"Pers #: {personnel}" in right_column:
                    found_pers = True
                    pers_column = 'right'
                    continue

                # If personnel is found, look for "0001-Basic Pay"
                if found_pers and "0001-Basic Pay" in line:
                    pay_pattern = r'0001-Basic Pay\s+([\d,]+\.\d{2})'

                    # Always prioritize the left column's match
                    left_match = re.search(pay_pattern, left_column)
                    if left_match:
                        basic_pay = left_match.group(1)
                        break  # Stop searching once found on the left

                    # If not found on the left, check the right
                    if pers_column == 'right':
                        right_match = re.search(pay_pattern, right_column)
                        if right_match:
                            basic_pay = right_match.group(1)

                    if basic_pay:
                        break  # Exit loop once the correct value is found

            if basic_pay:
                df.at[idx, 'Basic Pay'] = basic_pay
                print(f'Basic Pay for Pers # {personnel}: {basic_pay}')
            else:
                df.at[idx, 'Basic Pay'] = 'Not In Pay Roll'
                print(f'Basic Pay not found for Pers # {personnel}')

        df.to_excel(personnel_excel, index=False)

        # Load the workbook and select the active sheet
        workbook = load_workbook(personnel_excel)
        sheet = workbook.active

        # Define styles
        border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
        fill_style_header = PatternFill(start_color='FFCCFF', end_color='FFCCFF', fill_type='solid')  # Light purple
        fill_style_valid = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # Light yellow
        fill_style_row = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')  # Light gray for rows
        font_style = Font(name='Book Antiqua', bold=True, color='000000')  # Black text
        alignment_style = Alignment(wrap_text=True)

        # Apply styles to header row
        for col in range(1, len(df.columns) + 1):
            header_cell = sheet.cell(row=1, column=col)
            header_cell.fill = fill_style_header
            header_cell.font = font_style
            header_cell.border = border_style
            header_cell.alignment = alignment_style

        # Apply styles to data rows
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = border_style
                cell.alignment = alignment_style  # Wrap text for all cells

                # Fill entire row up to "Basic Pay" with background color
                if col < df.columns.get_loc('Basic Pay') + 1:
                    cell.fill = fill_style_row

                if col == df.columns.get_loc('Basic Pay') + 1 and cell.value == 'Not In Pay Roll':
                    cell.fill = fill_style_valid

        # Adjust column widths for better visibility
        for col in range(1, len(df.columns) + 1):
            column_letter = sheet.cell(row=1, column=col).column_letter
            max_length = max(len(str(cell.value)) for cell in sheet[column_letter])
            sheet.column_dimensions[column_letter].width = max_length + 2  # Adding padding

        # Save the workbook
        workbook.save(personnel_excel)

    except Exception as e:
        print(f'An error occurred: {str(e)}')


# Example usage
file_path = 'C:/Users/ACS/Downloads/SEPT-24.TXT'
personnel_excel_path = 'C:/Users/ACS/Pictures/PayRoll.xlsx'
extract_basic_pay(file_path, personnel_excel_path)
