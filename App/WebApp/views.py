import os
from datetime import datetime

import pandas as pd
from django.core.paginator import Paginator
from django.http import JsonResponse
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

"""
The code consists of Django views for managing user authentication, dashboard data, employee
transfer postings, leave applications, explanations, and related functionalities.

:param request: The code you provided is a Django application that includes views for managing user
authentication, dashboard display, employee transfer postings, leave applications, explanations, and
various data management functionalities. Here's a brief overview of the main views:
:return: The code provided contains views and functions for a Django web application related to
employee management, including functionalities for login, dashboard, transfer posting, leave
applications, explanations, and managing employee data. The code returns rendered HTML templates
with context data for the respective views, such as 'Dashboard.html', 'DispositionList.html',
'search.html', 'Zone.html', 'TransferPosting.html', 'ManageTransferPosting.html
"""
import mimetypes
import logging

from django.db.models import Sum, Count, Q
from django.shortcuts import render, HttpResponse, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required

from .Graph import transfer_posting_chart, get_employee_leave_data, get_employee_explanation_data, \
    getZoneRetirementList, get_age_range_count, get_zone_age_range_chart, get_retirement_year_count, get_zone_wise_count
from .Utitlities import (DesignationWiseList, getRetirementList, fetchAllDispositionList, getZoneWiseOfficialsList,
                         ZoneWiseStrength, ZoneDesignationWiseComparison, StrengthComparison, getAllEmpTransferPosting,
                         getAllEmpLeaveApplication, getAllEmpLeaveExplanation, is_admin, is_zone_admin, calculate_tax,
                         getAllBoardTransferPosting
                         )
from .models import DispositionList, TransferPosting, LeaveApplication, Explanation
from .tables import CountLeaveIndividuals_table, CountExplanationIndividuals_table, \
    CountTransferPostingIndividuals_table

logger = logging.getLogger(__name__)


def index(request):
    if request.user.is_authenticated:
        return redirect('Dashboard')
    return render(request, 'login.html')


def userLogin(request):
    if request.user.is_authenticated:
        return redirect('Dashboard')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            request.session['UserName'] = username
            return redirect('Dashboard')
        return render(request, 'login.html', {
            'title': 'Invalid',
            'icon': 'error',
            'message': 'Invalid Username or Password!'
        })

    return render(request, 'login.html', {
        'title': 'Invalid Method',
        'icon': 'error',
        'message': 'Method shall be POST rather than GET!'
    })


@login_required(login_url='userLogin')  # redirect when user is not logged in
def Dashboard(request):
    try:
        if not request.user.is_authenticated:
            return redirect('/')

        transfer_posting_summary = transfer_posting_chart(request)  # Graph Scripts
        leave_summary = get_employee_leave_data(request)  # Graph Scripts
        explanation_summary = get_employee_explanation_data(request)  # Graph Scripts
        zone_counts = getZoneRetirementList(request.user.userType, request)  # Graph Scripts
        age_range_count = get_age_range_count(request)  # Graph Scripts
        zone_age_ranges = get_zone_age_range_chart(request)  # Graph Scripts
        retirement_year_count = get_retirement_year_count(request)  # Graph Scripts
        zone_wise_count = get_zone_wise_count(request)

        label = "Employee yet to be Retired in the Year 2024, Regional Tax Office - II"
        Comparison = ZoneDesignationWiseComparison()
        results = DesignationWiseList(request.user.userType, request)
        employee_to_be_retired = getRetirementList(request.user.userType, request)  # table retired

        context = {
            'zones': list(zone_counts.keys()),  # Zones name Graph
            'counts': list(zone_counts.values()),  # No. of Employee Retired
            'retired': employee_to_be_retired,  # Retired Employee List
            'results': results,
            'Comparison': Comparison,
            'label': label,
            'leave_summary': leave_summary,  # Employee Leave Application Request
            'count_leave_individuals': CountLeaveIndividuals_table(request),  # Leave Summary Data In Form of Table
            'explanation_summary': explanation_summary,  # Graph Scripts
            'transfer_posting_summary': transfer_posting_summary,
            'CountExplanationIndividuals': CountExplanationIndividuals_table(request),  # Explanation Summary table Data
            'CountTransferPostingIndividuals': CountTransferPostingIndividuals_table(request),
            'age_range_count': age_range_count,
            'zone_age_ranges': zone_age_ranges,
            'retirement_year_count': retirement_year_count,
            'zone_wise_count': zone_wise_count

        }

        return render(request, 'Dashboard.html', context)
    except Exception as e:
        logger.error(f"Error in Dashboard view: {e}")
        return render(request, 'Dashboard.html', {'error': str(e)})


@login_required(login_url='userLogin')  # redirect when user is not logged in
def getDispositionList(request):
    try:
        search_query = request.GET.get('search', '')
        disposition_result = None  # Initialize variable

        if is_admin(request.user):
            if search_query:
                disposition_result = DispositionList.objects.filter(
                    (
                        Q(Personal_No__icontains=search_query) |
                        Q(CNIC_No__icontains=search_query) |
                        Q(Name__icontains=search_query) |
                        Q(ZONE__icontains=search_query) |
                        Q(Designation__icontains=search_query)
                ),
                status = 1  # Add this condition for filtering by status
                )
            else:
                disposition_result, error = fetchAllDispositionList(request)
                print(len(disposition_result))

        elif is_zone_admin(request.user):
            if search_query:
                # Filter by CNIC_No, Name, ZONE, Designation, and restrict by user's zone type
                disposition_result = DispositionList.objects.filter(
                    (
                            Q(Personal_No__icontains=search_query) |
                            Q(CNIC_No__icontains=search_query) |
                            Q(Name__icontains=search_query) |
                            Q(ZONE__icontains=search_query) |
                            Q(Designation__icontains=search_query)
                    ),
                    status=1  # Add this condition for filtering by status
                )
            else:
                disposition_result, error = fetchAllDispositionList(request)

        # Handle pagination
        if disposition_result is not None:
            page = request.GET.get('page')
            paginator = Paginator(disposition_result, 10)  # 10 items per page
            disposition_result = paginator.get_page(page)
            # Calculate the starting serial number for the current page
            start_serial_number = (disposition_result.number - 1) * paginator.per_page + 1

        return render(request, 'DispositionList.html', {'DispositionResult': disposition_result, 'search': search_query,
                                                        'start_serial_number': start_serial_number})

    except Exception as e:
        logger.error(f"Error in getDispositionList view: {e}")
        return HttpResponse("An error occurred.", status=500)


@login_required(login_url='userLogin')  # redirect when user is not logged in
def AddEditDisposition(request):
    try:
        if is_admin(request.user):
            empId = request.GET.get('empId')
            rowId = request.GET.get('rowId')
            if request.method == 'POST':
                empId = request.POST.get('emp_name')
                empStatus = request.POST.get('emp_status')
                employee = DispositionList.objects.get(id=empId,status=1)
                employee.status = empStatus
                employee.save()
                return HttpResponse(str("Yup !"))
            data = DispositionList.objects.filter(status=1)
            print(data)
            return render(request, 'AddEditDispositionList.html', {'result': data, 'empId' : empId,'rowId' : rowId})

    except Exception as e:
        logger.error(f"Error in getDispositionList view: {e}")
        return HttpResponse("An error occurred.", status=500)


@login_required(login_url='userLogin')  # redirect when user is not logged in
def Search(request):
    try:
        if is_admin(request.user):  # admin
            data = DispositionList.objects.values('id', 'Name', 'Designation', 'ZONE', 'CNIC_No', 'Date_of_Birth',
                                                  'Date_of_Entry_into_Govt_Service', 'Date_of_Retirement').filter(status=1)
        if is_zone_admin(request.user):
            data = DispositionList.objects.values('id', 'Name', 'Designation', 'ZONE', 'CNIC_No', 'Date_of_Birth',
                                                  'Date_of_Entry_into_Govt_Service', 'Date_of_Retirement').filter(
                ZONE=request.user.userType,status=1)

        if request.method == 'POST':
            # search_type = request.POST.get('type')
            search_value = request.POST.get('emp_name')
            leave_application = CountLeaveIndividuals_table(request, search_value)
            explanation_application = CountExplanationIndividuals_table(request, search_value)
            transfer_application = CountTransferPostingIndividuals_table(request, search_value)
            # filter_args = {'CNIC_No': search_value} if search_type == 'CNIC' else {'Personal_No': search_value}
            # result = DispositionList.objects.filter(id=search_value)
            result = DispositionList.objects.get(id=search_value)

            context = {
                'result': result,
                'data': data,
                'id': search_value,
                'leave_application': leave_application,
                'explanation_application': explanation_application,
                'transfer_application': transfer_application
            }
            return render(request, 'search.html', context)
        return render(request, 'search.html', {'data': data})
    except Exception as e:
        logger.error(f"Error in Search view: {e}")
        return HttpResponse("An error occurred.", status=500)


@login_required(login_url='userLogin')  # redirect when user is not logged in
def Zone(request):
    try:
        if is_admin(request.user):
            context = ZoneWiseStrength()
            if request.method == 'POST':
                search_zone = request.POST.get('Zone')
                results = getZoneWiseOfficialsList(search_zone)
                context.update({"results": results, 'zone': search_zone})
            return render(request, 'Zone.html', context)
        return redirect('/')
    except Exception as e:
        logger.error(f"Error in Zone view: {e}")
        return HttpResponse("An error occurred.", status=500)


@login_required(login_url='userLogin')  # redirect when user is not logged in
def Sanction_Strength(request):
    try:
        if is_admin(request.user):  # Checking if the user is an admin
            posts_data = [
                {"s_no": 1, "name": "Chief Commissioner-IR", "bs": 21, "sanctioned": 1, "working": 1, "vacancy": 0},
                {"s_no": 2, "name": "Commissioner-IR", "bs": 20, "sanctioned": 5, "working": 5, "vacancy": 0},
                {"s_no": 3, "name": "Additional Commissioner-IR", "bs": 19, "sanctioned": 11, "working": 4,
                 "vacancy": 7},
                {"s_no": 4, "name": "Cost Accountant", "bs": 19, "sanctioned": 1, "working": 1, "vacancy": 0},
                {"s_no": 5, "name": "Deputy/Assistant Commissioner-IR", "bs": "18/17", "sanctioned": 34, "working": 28,
                 "vacancy": 6},
                {"s_no": 6, "name": "Accounts Officer", "bs": 18, "sanctioned": 1, "working": 1, "vacancy": 0},
                {"s_no": 7, "name": "Deputy Director (MIS)", "bs": 18, "sanctioned": 1, "working": 0, "vacancy": 1},
                {"s_no": 8, "name": "Assistant Director Audit", "bs": 18, "sanctioned": 13, "working": 7, "vacancy": 6},
                {"s_no": 9, "name": "Treasury Officer", "bs": 17, "sanctioned": 1, "working": 0, "vacancy": 1},
                {"s_no": 10, "name": "Administrative Officer", "bs": 17, "sanctioned": 2, "working": 1, "vacancy": 1},
                {"s_no": 11, "name": "Assistant Director (MIS)", "bs": 17, "sanctioned": 4, "working": 4, "vacancy": 0},
                {"s_no": 12, "name": "Department Representative", "bs": 17, "sanctioned": 2, "working": 0,
                 "vacancy": 2},
                {"s_no": 13, "name": "MIS Officer", "bs": 16, "sanctioned": 10, "working": 12, "vacancy": -2},
                {"s_no": 14, "name": "Assistant Private Secretary", "bs": 16, "sanctioned": 11, "working": 9,
                 "vacancy": 2},
                {"s_no": 15, "name": "Inland Revenue Officer", "bs": 16, "sanctioned": 36, "working": 31,
                 "vacancy": 5},
                {"s_no": 16, "name": "Senior Auditor", "bs": 16, "sanctioned": 12, "working": 10, "vacancy": 2},
                {"s_no": 17, "name": "Inspector", "bs": 16, "sanctioned": 163, "working": 30, "vacancy": 133},
                {"s_no": 18, "name": "Superintendent", "bs": 16, "sanctioned": 11, "working": 0, "vacancy": 11},
                {"s_no": 19, "name": "Deputy Superintendent", "bs": 16, "sanctioned": 3, "working": 0, "vacancy": 3},
                {"s_no": 20, "name": "Office Superintendent", "bs": 16, "sanctioned": 26, "working": 17,
                 "vacancy": 9},
                {"s_no": 21, "name": "Total", "bs": "", "sanctioned": 348, "working": 164, "vacancy": 184},
                {"s_no": 22, "name": "Stenotypist", "bs": 14, "sanctioned": 42, "working": 16, "vacancy": 26},
                {"s_no": 23, "name": "Head Clerk", "bs": 14, "sanctioned": 3, "working": 15, "vacancy": -12},
                {"s_no": 24, "name": "Assistant", "bs": "15/16", "sanctioned": 2, "working": 1, "vacancy": 1},
                {"s_no": 25, "name": "Supervisor", "bs": 14, "sanctioned": 66, "working": 64, "vacancy": 2},
                {"s_no": 26, "name": "DEO *", "bs": 12, "sanctioned": 21, "working": 15, "vacancy": 6},
                {"s_no": 27, "name": "Library Assistant", "bs": 11, "sanctioned": 2, "working": 1, "vacancy": 1},
                {"s_no": 28, "name": "UDC", "bs": 11, "sanctioned": 162, "working": 72, "vacancy": 90},
                {"s_no": 29, "name": "LDC", "bs": 9, "sanctioned": 113, "working": 79, "vacancy": 34},
                {"s_no": 30, "name": "Hawaldar", "bs": 7, "sanctioned": 16, "working": 14, "vacancy": 2},
                {"s_no": 31, "name": "Telephone Operator/Telex Operator", "bs": 7, "sanctioned": 3, "working": 3,
                 "vacancy": 0},
                {"s_no": 32, "name": "Sepoy/Jamadar", "bs": 5, "sanctioned": 59, "working": 40, "vacancy": 19},
                {"s_no": 33, "name": "Dispatch Rider", "bs": 4, "sanctioned": 1, "working": 2, "vacancy": -1},
                {"s_no": 34, "name": "Motor Mechanic", "bs": 4, "sanctioned": 1, "working": 0, "vacancy": 1},
                {"s_no": 35, "name": "Duplicating Machine Operator (DMO)", "bs": 4, "sanctioned": 1, "working": 0,
                 "vacancy": 1},
                {"s_no": 36, "name": "Driver", "bs": 4, "sanctioned": 29, "working": 25, "vacancy": 4},
                {"s_no": 37, "name": "Daftari", "bs": 2, "sanctioned": 41, "working": 28, "vacancy": 13},
                {"s_no": 38, "name": "Record Sorter", "bs": 2, "sanctioned": 5, "working": 3, "vacancy": 2},
                {"s_no": 39, "name": "Qasid", "bs": 2, "sanctioned": 3, "working": 0, "vacancy": 3},
                {"s_no": 40, "name": "Bailiff", "bs": 1, "sanctioned": 18, "working": 12, "vacancy": 6},
                {"s_no": 41, "name": "Chowkidar", "bs": 1, "sanctioned": 2, "working": 1, "vacancy": 1},
                {"s_no": 42, "name": "Mali", "bs": 1, "sanctioned": 6, "working": 8, "vacancy": -2},
                {"s_no": 43, "name": "Farash", "bs": 1, "sanctioned": 8, "working": 2, "vacancy": 6},
                {"s_no": 44, "name": "Naib Qasid", "bs": 1, "sanctioned": 120, "working": 49, "vacancy": 71},
                {"s_no": 45, "name": "Dresser", "bs": 1, "sanctioned": 1, "working": 0, "vacancy": 1},
                {"s_no": 46, "name": "Notice Server *", "bs": 1, "sanctioned": 34, "working": 0, "vacancy": 34},
                {"s_no": 47, "name": "Water Carrier", "bs": 1, "sanctioned": 1, "working": 0, "vacancy": 1},
                {"s_no": 48, "name": "Sanitary Worker", "bs": 1, "sanctioned": 4, "working": 0, "vacancy": 4},
                {"s_no": 49, "name": "Armed Guard", "bs": 1, "sanctioned": 14, "working": 7, "vacancy": 7},
                {"s_no": 50, "name": "Total BS-01 to 16", "bs": "", "sanctioned": 778, "working": 458, "vacancy": 320},
                {"s_no": 51, "name": "Grand Total", "bs": "", "sanctioned": 1126, "working": 622, "vacancy": 504}
            ]
            paginator = Paginator(posts_data, 10)  # Show 10 posts per page
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            return render(request, 'Sanction_Strenght.html', {'page_obj': page_obj})
            # return render(request, 'Sanction_Strenght.html', {'posts_data': posts_data})
        else:
            return HttpResponse("You do not have permission to view this page.")
    except Exception as e:
        logger.error(f"Error in Sanction & Strength view: {e}")
        return HttpResponse(f"An error occurred: {str(e)}")


@login_required(login_url='userLogin')  # redirect when user is not logged in
def EmployeeTransferPosting(request):  # Transfer Posting Form
    try:
        row = {}
        empId = request.GET.get('empId', '')
        rowId = request.GET.get('rowId', '')
        opType = request.GET.get('type', '')
        userType = request.GET.get('userType', '')
        # Determine if the user is an admin or zone admin and filter data accordingly
        if is_admin(request.user):
            data = DispositionList.objects.values('id', 'Name', 'Designation', 'ZONE').filter(status=1)
        if is_zone_admin(request.user):
            data = DispositionList.objects.values('id', 'Name', 'Designation', 'ZONE').filter(
                ZONE=request.user.userType,status=1)
        # super admin will edit the record
        if empId and rowId and is_admin(request.user):
            print('admin / CCIR')
            postingRow = TransferPosting.objects.get(id=rowId, employee_id=empId)
            if postingRow:
                row = {
                    'old_zone': postingRow.old_zone or '',
                    'new_zone': postingRow.new_zone or '',
                    'chief_order_number': postingRow.chief_order_number or '',
                    'chief_transfer_date': postingRow.chief_transfer_date or '',
                    'chief_reason_for_transfer': postingRow.chief_reason_for_transfer or '',
                    'chief_order_approved_by': postingRow.chief_order_approved_by or ''

                }  # # f
            else:
                # Handle the case where both empId and rowId are None
                row = {
                    'old_zone': '',
                    'new_zone': '',
                    'chief_order_number': '',
                    'chief_transfer_date': '',
                    'chief_reason_for_transfer': '',
                    'chief_order_approved_by': ''

                }
        # fetch row from database zone admin in case of edit record
        elif empId and rowId and is_zone_admin(request.user):
            print('zone admin edit record')
            employee = DispositionList.objects.get(id=empId)
            postingRow = TransferPosting.objects.get(id=rowId, employee_id=empId, zone_type=request.user.userType)
            if postingRow:
                row = {
                    'zone_current_unit': postingRow.old_unit or '',
                    'zone_new_unit': postingRow.new_unit or '',
                    'zone_order_number': postingRow.zone_order_number or '',
                    'zone_transfer_date': postingRow.zone_transfer_date or '',
                    'zone_transfer_reason': postingRow.zone_reason_for_transfer or '',
                    'zone_order_approved_by': postingRow.zone_order_approved_by or ''

                }  # # f
            else:
                # Handle the case where both empId and rowId are None
                row = {
                    'zone_current_unit': '',
                    'zone_new_unit': '',
                    'zone_order_number': '',
                    'zone_transfer_date': '',
                    'zone_transfer_reason': '',
                    'zone_order_approved_by': ''

                }
            # You can add additional logic here, like logging an error or returning a response

        if request.method == 'POST':
            emp_name = request.POST.get('emp_name') or request.POST.get('hd_emp')
            image = request.FILES.get('image')
            hd_rowId = request.POST.get('hd_rowId')
            hd_type = request.POST.get('hd_type')

            # Validate uploaded file
            if not image:
                return HttpResponse("No file uploaded.", status=400)

            file_type = mimetypes.guess_type(image.name)[0]
            if not file_type or (not file_type.startswith('image') and file_type != 'application/pdf'):
                return HttpResponse("Uploaded file is not a valid image or PDF.", status=400)

            # Retrieve the employee object
            employee = DispositionList.objects.get(id=emp_name)

            if is_admin(request.user) and hd_type:
                new_zone = request.POST.get('new_zone')
                employee.ZONE = new_zone
                employee.save()
                print('admin edit record / ccir')
                transfer_posting = TransferPosting.objects.get(employee=employee, id=hd_rowId)
                transfer_posting.old_zone = request.POST.get('old_zone')
                transfer_posting.new_zone = request.POST.get('new_zone')
                transfer_posting.chief_transfer_date = int(request.POST.get('order_number'))
                transfer_order_date_str = request.POST.get('transfer_order_date')
                try:
                    transfer_order_date = datetime.fromisoformat(transfer_order_date_str)
                except ValueError:
                    transfer_order_date = None

                transfer_posting.chief_transfer_date = transfer_order_date
                transfer_posting.chief_reason_for_transfer = request.POST.get('transfer_reason')
                transfer_posting.chief_order_approved_by = request.POST.get('order_approved_by')
                transfer_posting.chief_transfer_document = image
                transfer_posting.zone_type = new_zone

                transfer_posting.save()

            elif is_admin(request.user):  # ccir will create record
                print('admin create record')
                # Handling for superuser (admin)
                new_zone = request.POST.get('new_zone')
                employee.ZONE = new_zone
                employee.save()

                transfer_posting = TransferPosting(
                    employee=employee,
                    old_zone=request.POST.get('old_zone'),
                    new_zone=new_zone,
                    chief_order_number=request.POST.get('order_number'),
                    chief_transfer_date=request.POST.get('transfer_order_date'),
                    chief_reason_for_transfer=request.POST.get('transfer_reason'),
                    chief_order_approved_by=request.POST.get('order_approved_by'),
                    chief_transfer_document=image,
                    zone_type=new_zone
                )

            elif is_zone_admin(request.user) and hd_type and hd_rowId:  # after CCIR ZOne will assigned Unit
                print("editing zone record employees.......")
                # Handling for zone admin
                transfer_posting = TransferPosting.objects.get(employee=employee, zone_type=request.user.userType,
                                                               id=hd_rowId)
                transfer_posting.old_unit = request.POST.get('zone_prev_unit')
                transfer_posting.new_unit = request.POST.get('zone_new_unit')
                transfer_posting.zone_range = request.POST.get('range')
                transfer_posting.zone_order_number = int(request.POST.get('order_number'))
                transfer_order_date_str = request.POST.get('transfer_order_date')
                try:
                    transfer_order_date = datetime.fromisoformat(transfer_order_date_str)
                except ValueError:
                    transfer_order_date = None

                transfer_posting.zone_transfer_date = transfer_order_date
                transfer_posting.zone_reason_for_transfer = request.POST.get('transfer_reason')
                transfer_posting.zone_order_approved_by = request.POST.get('order_approved_by')
                transfer_posting.zone_transfer_document = image

            else:  # in case no emp Id and row Id creating transfer posting by zone,Internally Posting Employee by ZONE
                print('create transfer posting')
                # Fetch the latest TransferPosting record for this employee if it's an internal transfer
                emp_zone = DispositionList.objects.filter(id=emp_name).first()
                previous_transfer = TransferPosting.objects.filter(employee=employee).order_by('-created_at').first()
                transfer_posting = TransferPosting(
                    employee=employee,
                    old_unit=request.POST.get('zone_prev_unit'),
                    new_unit=request.POST.get('zone_new_unit'),
                    zone_range=request.POST.get('range'),
                    zone_order_number=request.POST.get('order_number'),
                    zone_transfer_date=request.POST.get('transfer_order_date'),
                    zone_reason_for_transfer=request.POST.get('transfer_reason'),
                    zone_order_approved_by=request.POST.get('order_approved_by'),
                    zone_transfer_document=image,
                    zone_type=request.user.userType,
                    old_zone=emp_zone.ZONE if previous_transfer else '-',
                    new_zone=emp_zone.ZONE if previous_transfer else '-',
                    chief_order_number=previous_transfer.chief_order_number if previous_transfer else 0,
                    chief_transfer_date=previous_transfer.chief_transfer_date if previous_transfer else None,
                    chief_reason_for_transfer=previous_transfer.chief_reason_for_transfer if previous_transfer else '',
                    chief_order_approved_by=previous_transfer.chief_order_approved_by if previous_transfer else '',
                    chief_transfer_document=previous_transfer.chief_transfer_document if previous_transfer else '',

                )

            transfer_posting.save()

            return render(request, 'TransferPosting.html', {
                'title': 'Transfer Posting!',
                'icon': 'success',
                'message': 'Data Inserted Successfully!',
                'data': data,
            })

        return render(request, 'TransferPosting.html',
                      {'data': data, 'empId': empId, 'row': row, 'rowId': rowId, 'type': opType})
    except Exception as e:
        logger.error(f"Error in EmployeeTransferPosting view: {e}")
        return HttpResponse(f"An error occurred: {str(e)}", status=500)


@login_required(login_url='userLogin')  # Template Data
def ManageEmployeeTransferPosting(request):
    try:
        transfer_records = getAllEmpTransferPosting(request.user.is_superuser, request.user.userType)

        for item in transfer_records:
            zone_transfer_document = item.get('zone_transfer_document')
            chief_transfer_document = item.get('chief_transfer_document')
            mime_type = mimetypes.guess_type(zone_transfer_document)[0] if zone_transfer_document else None
            if mime_type == 'application/pdf':
                item['is_pdf'] = True
            elif mime_type and mime_type.startswith('image'):
                item['is_image'] = True

            mime_type = mimetypes.guess_type(chief_transfer_document)[0] if chief_transfer_document else None
            if mime_type == 'application/pdf':
                item['chief_is_pdf'] = True
            elif mime_type and mime_type.startswith('image'):
                item['chief_is_image'] = True

        page = request.GET.get('page')
        paginator = Paginator(transfer_records, 12)  # 10 items per page
        result = paginator.get_page(page)
        # Calculate the starting serial number for the current page
        start_serial_number = (result.number - 1) * paginator.per_page + 1

        return render(request, 'ManageTransferPosting.html', {
            'transfer_records': result, 'start_serial_number': start_serial_number
        })
    except Exception as e:
        logger.error(f"Error in ManageEmployeeTransferPosting view: {e}")
        return HttpResponse("An error occurred.", status=500)


@login_required(login_url='userLogin')  # Template Data
def ManageBoard(request):
    try:
        print("board")
        transfer_records = getAllBoardTransferPosting(request.user.is_superuser, request.user.userType)

        for item in transfer_records:
            chief_transfer_document = item.get('chief_transfer_document')

        page = request.GET.get('page')
        paginator = Paginator(transfer_records, 12)  # 10 items per page
        result = paginator.get_page(page)
        start_serial_number = (result.number - 1) * paginator.per_page + 1

        return render(request, 'Board.html', {
            'transfer_records': result, 'start_serial_number': start_serial_number
        })
    except Exception as e:
        logger.error(f"Error in ManageEmployeeTransferPosting view: {e}")
        return HttpResponse("An error occurred.", status=500)


@login_required(login_url='userLogin')  # Redirect when user is not logged in
def submitLeaveApplication(request):  # Leave Submission Form
    try:
        empId = request.GET.get('empId', '')
        rowId = request.GET.get('rowId', '')
        row = {}
        context = {}

        # Fetch data based on user role
        if is_admin(request.user):
            data = DispositionList.objects.values('id', 'Name', 'Designation', 'ZONE').filter(status=1)
        elif is_zone_admin(request.user):
            data = DispositionList.objects.values('id', 'Name', 'Designation', 'ZONE').filter(
                ZONE=request.user.userType,status=1)
        else:
            data = DispositionList.objects.none()  # No data if not admin or zone admin

        # Determine if the user is an admin or a zone admin and fetch the relevant row
        if rowId and empId:
            query_params = {'id': rowId, 'employee_id': empId}
            if is_zone_admin(request.user):
                query_params['zone_type'] = request.user.userType

            row = LeaveApplication.objects.filter(**query_params).first()

        if request.method == 'POST':
            hd_rowId = request.POST.get('hd_rowId')
            employee_name = request.POST.get('emp_name') or request.POST.get('hd_emp')
            leave_type = request.POST.get('leave_type')
            leave_start_date = request.POST.get('leave_start_date')
            leave_end_date = request.POST.get('leave_end_date')
            reason = request.POST.get('reason')
            leave_document = request.FILES.get('leave_document')
            zone_type = request.POST.get('zone_type') or request.POST.get('hd_zone_type')

            # Validate and process leave dates
            try:
                start_date = datetime.strptime(leave_start_date, '%Y-%m-%d')
                end_date = datetime.strptime(leave_end_date, '%Y-%m-%d')
                leave_duration = (end_date - start_date).days + 1
            except ValueError:
                context.update({'message': 'Invalid date format.', 'alert_type': 'error'})
                return render(request, 'LeaveApplication.html', context)

            # Calculate the number of days of leave already taken in the current year
            current_year = datetime.now().year
            leave_taken = LeaveApplication.objects.filter(
                employee=employee_name,
                leave_type=leave_type,
                leave_start_date__year=current_year
            ).aggregate(Sum('days_granted'))['days_granted__sum'] or 0

            # Check if updating an existing record

            if leave_type == 'Casual Leave' and leave_duration > 20:
                context.update({'message': 'Casual Leave cannot exceed 20 days per year.', 'alert_type': 'error'})
                return render(request, 'LeaveApplication.html', context)
            elif leave_type == 'Earned Leave' and leave_duration > 48:
                context.update({'message': 'Earned Leave cannot exceed 48 days per year.', 'alert_type': 'error'})
                return render(request, 'LeaveApplication.html', context)

            # Save or update the leave application
            if hd_rowId:
                # Update existing record
                query_params = {'id': hd_rowId, 'employee_id': employee_name}
                if is_zone_admin(request.user):
                    query_params['zone_type'] = request.user.userType
                elif is_admin(request.user):
                    query_params['zone_type'] = request.POST.get('hd_zone_type')

                row = LeaveApplication.objects.filter(**query_params).first()

                if row:
                    row.leave_type = leave_type
                    row.leave_start_date = leave_start_date
                    row.leave_end_date = leave_end_date
                    row.leave_document = leave_document
                    row.reason = reason
                    row.days_granted = leave_duration
                    row.save()
                    context.update(
                        {'message': 'Record Updated Successfully.', 'icon': 'success', 'empId': employee_name})
                else:
                    context.update({'message': 'Record not found.', 'icon': 'error'})
            else:
                # Create new record
                LeaveApplication.objects.create(
                    employee=DispositionList.objects.get(id=employee_name),
                    leave_type=leave_type,
                    leave_start_date=leave_start_date,
                    leave_end_date=leave_end_date,
                    leave_document=leave_document,
                    reason=reason,
                    days_granted=leave_duration,
                    zone_type=request.user.userType if is_zone_admin(request.user) else zone_type
                )
                context.update(
                    {'message': 'Your leave application has been submitted successfully.', 'icon': 'success',
                     'empId': employee_name})

        context.update({'data': data, 'rowId': rowId, 'empId': empId, 'row': row})
        return render(request, 'LeaveApplication.html', context)

    except Exception as e:
        return HttpResponse(str(e), status=500)


@login_required(login_url='userLogin')  # redirect when user is not logged in
def ManageEmployeeLeaveApplication(request):  # Data Template Leave
    try:
        leave_application = getAllEmpLeaveApplication(request.user.is_superuser, request.user.userType)
        for item in leave_application:
            leave_document = item.get('leave_document')
            mime_type = mimetypes.guess_type(leave_document)[0] if leave_document else None
            if mime_type == 'application/pdf':
                item['is_pdf'] = True
            elif mime_type and mime_type.startswith('image'):
                item['is_image'] = True

        return render(request, 'ManageLeaveApplication.html', {'leave_application': leave_application})
    except Exception as e:
        logger.error(f"Error in ManageLeaveApplication view: {e}")
        return HttpResponse("An error occurred.", status=500)


@login_required(login_url='userLogin')  # Redirect when user is not logged in
def EmployeeExplanation(request):  # Explanation Submission Form
    try:
        empId = request.GET.get('empId', '')
        rowId = request.GET.get('rowId', '')
        row = {}
        context = {}

        # Determine if the user is an admin or a zone admin and fetch the relevant row
        if rowId and empId:
            query_params = {'id': rowId, 'employee_id': empId}
            if is_zone_admin(request.user):
                query_params['zone_type'] = request.user.userType

            row = Explanation.objects.filter(**query_params).first()

        # Fetch data based on user role
        user_zone = request.user.userType if is_zone_admin(request.user) else None
        data = DispositionList.objects.values('id', 'Name', 'Designation', 'ZONE').filter(status=1)
        if user_zone:
            data = data.filter(ZONE=user_zone)

        if request.method == 'POST':
            rowId = request.POST.get('hd_rowId')
            hd_emp = request.POST.get('emp_name') or request.POST.get('hd_emp')
            employee = DispositionList.objects.get(id=hd_emp)
            exp_type = request.POST.get('exp_type')
            exp_issue_date = datetime.strptime(request.POST.get('exp_issue_date'), "%Y-%m-%d") if request.POST.get(
                'exp_issue_date') else None
            exp_reply_date = datetime.strptime(request.POST.get('exp_reply_date'), "%Y-%m-%d") if request.POST.get(
                'exp_reply_date') else None
            exp_document = request.FILES.get('exp_document')

            # Validate file upload
            if not exp_document:
                return HttpResponse("No file uploaded.", status=400)

            file_type = mimetypes.guess_type(exp_document.name)[0]
            if not file_type or (not file_type.startswith('image') and file_type != 'application/pdf'):
                return HttpResponse("Uploaded file is not a valid image or PDF.", status=400)

            # Update or create Explanation record based on rowId
            if rowId:
                query_params = {'id': rowId, 'employee_id': hd_emp}
                if is_zone_admin(request.user):
                    query_params['zone_type'] = request.user.userType
                elif is_admin(request.user):
                    query_params['zone_type'] = request.POST.get('hd_zone_type')

                row = Explanation.objects.filter(**query_params).first()
                if row:
                    row.exp_type = exp_type
                    row.exp_issue_date = exp_issue_date
                    row.exp_reply_date = exp_reply_date
                    row.exp_document = exp_document
                    row.save()
                    context.update({'alert_message': "Record Updated Successfully.", 'alert_type': 'success'})
                else:
                    context.update({'alert_message': "Record not found.", 'alert_type': 'error'})
            else:
                zone_type = request.user.userType if is_zone_admin(request.user) else request.POST.get('zone_type')
                Explanation.objects.create(
                    employee=employee,
                    exp_type=exp_type,
                    exp_issue_date=exp_issue_date,
                    exp_reply_date=exp_reply_date,
                    zone_type=zone_type,
                    exp_document=exp_document
                )
                context.update({'alert_message': "Record Created Successfully.", 'alert_type': 'success'})

        context.update({'data': data, 'rowId': rowId, 'empId': empId, 'row': row})
        return render(request, 'Explanation.html', context)
    except Exception as e:
        return HttpResponse(str(e), status=500)


@login_required(login_url='userLogin')  # redirect when user is not logged in
def ManageEmployeeExplanation(request):  # Data Template
    try:
        employee_explanation = getAllEmpLeaveExplanation(request.user.is_superuser, request.user.userType)
        for item in employee_explanation:
            exp_document = item.get('exp_document')
            mime_type = mimetypes.guess_type(exp_document)[0] if exp_document else None
            if mime_type == 'application/pdf':
                item['is_pdf'] = True
            elif mime_type and mime_type.startswith('image'):
                item['is_image'] = True

        return render(request, 'ManageExplanation.html', {'exp_document': employee_explanation})
    except Exception as e:
        logger.error(f"Error in Manage Explanation view: {e}")
        return HttpResponse("An error occurred.", status=500)


@login_required(login_url='userLogin')  # redirect when user is not logged in
def Strength(request):
    try:
        if is_admin(request.user):
            final_data = StrengthComparison(request.user.userType, None)
            # return HttpResponse(str(final_data))
            return render(request, 'strength.html', {
                'data': final_data,
                'Comparison': ZoneDesignationWiseComparison(),
            })
        return redirect('userLogin')
    except Exception as e:
        logger.error(f"Error in Strength view: {e}")
        return render(request, 'strength.html', {'error_message': str(e)})


def Logout(request):
    logout(request)
    return redirect('/')


def verify(request):
    try:
        # This function remains a GET request
        def read_txt_file(file_path):
            with open(file_path, 'r') as file:
                data = file.readlines()  # Read each line of the text file
            return data

        def check_record_in_db(record_id):
            # Check if the record exists in the database
            return DispositionList.objects.filter(Personal_No=record_id).exists()

        def check_records(file_path):
            records = read_txt_file(file_path)
            all_records = []

            for record in records:
                record_id = record.strip()  # Process the record data
                if check_record_in_db(record_id):
                    all_records.append(f"{record_id} - exists")
                else:
                    all_records.append(f"{record_id} - does not exist")

            # Save all records (both existing and non-existing) to a single file
            output_file_path = os.path.join('C:/Users/ACS/Documents', 'records_verification.txt')
            with open(output_file_path, 'w') as output_file:
                for record in all_records:
                    output_file.write(f"{record}\n")

            return output_file_path

        # Hardcoded file path, as in your original code
        file_path = 'C:/Users/ACS/Documents/personnel.txt'

        # Call the check_records function to check records and save them
        output_file_path = check_records(file_path)

        # Return the file path of saved results
        return JsonResponse({
            'status': 'success',
            'message': 'Records verified successfully.',
            'output_file_path': output_file_path
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })


def TaxSlab(request):
    try:
        if request.method == 'POST':
            income_type = request.POST.get('income_type')  # 'monthly' or 'yearly'
            income_amount = int(request.POST.get('income_amount'))  # Either monthly or yearly salary based on user
            taxpayer_type = request.POST.get('taxpayer_type')  # Either monthly or yearly salary based on user type
            excel_file = request.FILES.get('excel_file')

            print(taxpayer_type)

            # If monthly income, convert it to yearly income
            if income_type == 'Monthly':
                yearly_income = income_amount * 12
                print(yearly_income)
            else:
                yearly_income = income_amount  # Already yearly income
                print(yearly_income)

            # Define the tax brackets for 2023 and 2024
            tax_brackets_2023_2024_salaried = {
                (0, 600000): (0, 0),
                (600001, 1200000): (0.025, 600000),
                (1200001, 2400000): (0.125, 15000),
                (2400001, 3600000): (0.225, 165000),
                (3600001, 6000000): (0.275, 435000),
                (6000001, float('inf')): (0.35, 1095000)
            }

            tax_brackets_2024_2025_salaried = {
                (0, 600000): (0, 0),
                (600001, 1200000): (0.05, 600000),
                (1200001, 2200000): (0.15, 30000),
                (2200001, 3200000): (0.25, 180000),
                (3200001, 4100000): (0.30, 430000),
                (4100001, float('inf')): (0.35, 700000)
            }
            tax_brackets_business_2023_2024 = {
                (0, 600000): (0, 0),
                (600001, 800000): (0.075, 600000),  # Slightly higher than salaried
                (800001, 1200000): (0.15, 15000),
                (1200001, 2400000): (0.2, 75000),
                (2400001, 3000000): (0.25, 315000),  # Higher bracket percentages
                (3000001, 4000000): (0.3, 465000),  # Adjusted for higher business earnings
                (4000001, float('inf')): (0.35, 765000)

            }
            tax_brackets_business_2024_2025 = {
                (0, 600000): (0, 0),
                (600001, 1200000): (0.15, 600000),  # Slightly higher than salaried
                (1200001, 1600000): (0.2, 90000),
                (1600001, 3200000): (0.3, 170000),
                (3200001, 5600000): (0.4, 650000),  # Higher bracket percentages
                (5600001, float('inf')): (0.45, 1610000)
            }

            apply_surcharge_2023 = False
            apply_surcharge_2024 = True

            # Calculate taxes for both 2023 and 2024 based on yearly income
            if taxpayer_type == 'salaried':
                tax_2023 = calculate_tax(yearly_income, tax_brackets_2023_2024_salaried, apply_surcharge_2023)
                tax_2024 = calculate_tax(yearly_income, tax_brackets_2024_2025_salaried, apply_surcharge_2024)
            if taxpayer_type == 'business':
                tax_2023 = calculate_tax(yearly_income, tax_brackets_business_2023_2024, apply_surcharge_2023)
                tax_2024 = calculate_tax(yearly_income, tax_brackets_business_2024_2025, apply_surcharge_2024)

            # Calculate the percentage tax and growth between 2023 and 2024
            if yearly_income == 0:
                tax_2023_percentage = 0
                tax_2024_percentage = 0
            else:
                tax_2023_percentage = (tax_2023['total_tax'] / yearly_income) * 100 if tax_2023['total_tax'] > 0 else 0
                tax_2024_percentage = (tax_2024['total_tax'] / yearly_income) * 100 if tax_2024['total_tax'] > 0 else 0

                if tax_2023_percentage > 0 and tax_2024_percentage > 0:
                    growth_percentage = ((tax_2024_percentage - tax_2023_percentage) / tax_2023_percentage) * 100
                    growth_percentage = round(growth_percentage, 2)
                else:
                    growth_percentage = 0

            # Prepare the context for the template
            context = {
                'tax_2023_year': '2023 - 2024',
                'tax_2024_year': '2024 - 2025',
                'tax_2023': tax_2023,
                'tax_2024': tax_2024,
                'tax_2023_percentage': tax_2023_percentage,
                'tax_2024_percentage': tax_2024_percentage,
                'yearly_income': yearly_income,
                'monthly_income': income_amount if income_type == 'Monthly' else yearly_income,
                'growth_percentage': growth_percentage,
                'income_type': income_type,
                'taxpayer_type': taxpayer_type
            }
            return render(request, 'TaxSlab.html', context)

        # Render an empty form when the page is loaded initially
        context = {
            'tax_2023_year': '2023 - 2024',
            'tax_2024_year': '2024 - 2025',
            'tax_2023': '',
            'tax_2024': '',
            'tax_2023_percentage': '',
            'tax_2024_percentage': '',
            'yearly_income': '',
            'growth_percentage': '',
            'income_type': '',
            'taxpayer_type': ''

        }
        return render(request, 'TaxSlab.html', context)

    except Exception as e:
        print(str(e))
        return HttpResponse(str(e))


def ExcelImport(request):
    try:
        if request.method == 'POST':
            # Tax brackets (adjust as per actual requirements)

            excel_file = request.FILES.get('excel_file')

            if not excel_file:
                return HttpResponse("Please upload a file")

            try:
                # Read the uploaded Excel file into a DataFrame
                df = pd.read_excel(excel_file)
            except Exception as e:
                return HttpResponse(f"Error reading file: {str(e)}")

            # Ensure 'MONTHLY_SALARY' column exists
            if 'MONTHLY_SALARY' not in df.columns:
                return HttpResponse("Excel file must contain a 'MONTHLY_SALARY' column.")

            # Add required new columns
            df['Annual_Salary'] = df['MONTHLY_SALARY'] * 12
            tax_results = []
            TAX_BRACKETS = {
                (0, 600000): (0, 0),
                (600001, 1200000): (0.05, 600000),
                (1200001, 2200000): (0.15, 30000),
                (2200001, 3200000): (0.25, 180000),
                (3200001, 4100000): (0.30, 430000),
                (4100001, float('inf')): (0.35, 700000)
            }

            for salary in df['Annual_Salary']:
                tax_info = calculate_tax(salary, TAX_BRACKETS, apply_surcharge=True)
                tax_results.append(tax_info)

            # Append calculated columns to DataFrame
            df['Annual_Tax'] = [result['total_tax'] for result in tax_results]
            df['4AB'] = [result['surcharge'] for result in tax_results]
            df['Annual Normal + 4AB'] = [result['total_tax_with_surcharge'] for result in tax_results]
            df['Monthly Normal + 4AB'] = [result['per_month'] for result in tax_results]
            # Convert TAX_DEDUCTED to numeric and calculate Difference
            df['TAX_DEDUCTED'] = pd.to_numeric(df['TAX_DEDUCTED'], errors='coerce').fillna(0)
            df['Difference'] = [
                result['per_month'] - tax_deducted
                for result, tax_deducted in zip(tax_results, df['TAX_DEDUCTED'])
            ]

            # Create an Excel file with formatting using openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = "Tax Calculations"

            # Add headers with formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            # Write headers
            for col_num, header in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Write data rows with borders
            for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Apply number formatting
                    if ws.cell(row=1, column=col_idx).value in ['MONTHLY_SALARY', 'Annual_Salary', 'Annual_Tax', '4AB',
                                                                'Annual Normal + 4AB', 'Monthly Normal + 4AB',
                                                                'Difference', 'TAX_DEDUCTED']:
                        cell.number_format = '#,##0'  # Currency/Decimal format
                    elif ws.cell(row=1, column=col_idx).value == 'Percentage':
                        cell.number_format = '0.00%'  # Example if you have percentage columns

            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2

            # Export to Excel file
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="calculated_taxes.xlsx"'
            wb.save(response)

            return response

        return render(request, 'ImportExcel.html')
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}")
